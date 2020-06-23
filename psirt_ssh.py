#
# to get client_id and cleint_secret please refer to
# https://apiconsole.cisco.com/docs/read/overview/Platform_Introduction
# and
# https://developer.cisco.com/psirt/
#

import json
import urllib
import requests
import logging
import argparse
import re
import getpass
import config
from urllib3.exceptions import InsecureRequestWarning
from netmiko.ssh_autodetect import SSHDetect
from netmiko.ssh_dispatcher import ConnectHandler
from netmiko.ssh_exception import NetmikoAuthenticationException
from netmiko.ssh_exception import NetmikoTimeoutException
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import GREEN


TOKEN_URL = 'https://cloudsso.cisco.com/as/token.oauth2'
REST_URL = 'https://api.cisco.com/security/advisories'


def ssh_version(host: str, username: str, password: str):
    """ Get Device Type and Version using SSH Netmiko
    :param host: Hostname or IP adress of the router/switch
    :param username: Username
    :param password: Password
    :return: list [device, version]

    """
    command = 'show version'
    remote_device = {'device_type': 'cisco_ios',
                     'host': host,
                     'username': username,
                     'password': password}

    VERSION_MAPPER = {
        'ios': {
            'device_pattern': [
                r'Cisco IOS Software', r'Cisco Internetwork Operating System Software'],
            'version_pattern': r'Cisco.IOS.Software.*Version\s+([^,\s]+).+'
        },

        'nxos': {
            'device_pattern': [r'Cisco Nexus Operating System', r'NX-OS'],
            'version_pattern': r'.*version (\d+.+)$'

        },
        'iosxe': {
            'device_pattern': [
                r'Cisco IOS XE Software', r'IOS-XE'],
            'version_pattern': 'Cisco.IOS.XE.Software.*Version\s+([^,\s]+)[.|\s]+'
        }

    }

    try:
        conn = ConnectHandler(**remote_device)
        output = conn.send_command(command)
        logging.debug(output)
    except (NetmikoAuthenticationException, NetmikoTimeoutException) as e:
        logging.error(e)
        exit(1)

    for device_type, patterns in VERSION_MAPPER.items():
        for pattern in patterns['device_pattern']:
            match = re.search(pattern, output, flags=re.I)
            if match:
                version = re.findall(
                    patterns['version_pattern'], output, re.MULTILINE)
                if version:
                    return [device_type, version[0]]


def nxos_rest_version(host: str, username: str, password: str):
    """ Get Device Type and Version using NXOS REST API
    :param host: Hostname or IP adress of the router/switch
    :param username: Username
    :param password: Password
    :return: string with version

    """
    requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)
    s = requests.Session()
    s.auth = (username, password)
    s.verify = False
    s.headers.update({'Content-Type': 'application/json'})
    url = 'https://' + host + '/ins'
    data = {'ins_api': {'chunk': '0',
                        'version': '1.0',
                        'sid': '1',
                        'input': 'show version',
                        'type': 'cli_show',
                        'output_format': 'json'}
            }
    try:
        ret = s.request('post', url, data=json.dumps(data))
    except requests.exceptions.ConnectionError as e:
        exit(1)
        logging.error('Error: %s', e)
    except requests.exceptions.Timeout:
        logging.error('Could not login due to timeout')
        exit(1)
    if ret.ok:
        output = ret.json()['ins_api']['outputs']['output']['body']
        if 'nxos_ver_str' in output:
            return output['nxos_ver_str']
        elif 'sys_ver_str' in output:
            return output['sys_ver_str']
    elif ret.status_code == 401:
        logging.error('Error: User Authentication Failure')


def get_auth_token(client_id: str, client_secret: str):
    """ GET authentcation token  
    :param client_id: Client ID from API portal
    :param client_secret: Client Secret from API portal
    :return: string with token

    """
    ret = requests.post(TOKEN_URL,
                        params={'client_id': client_id,
                                'client_secret': client_secret},
                        data={'grant_type': 'client_credentials'}
                        )
    if ret.ok:
        return ret.json()['access_token']
    else:
        logging.error(
            'Error: Cannot get token. Request Status Code: ', ret.status_code)
        exit(1)


def get_psirt(token: str, os_type: str, version: str):
    """ Get PSIRT's for device
    :param token: OpenVulnAPI access token
    :param os_type: device OS type: nxos, ios, iosxe
    :param version: device version. e.g. 16.06.03 

    """
    url = 'https://api.cisco.com/security/advisories/'
    s = requests.Session()
    params = {'version': version}
    s.headers.update({'Authorization': 'Bearer ' + token})
    logging.info('Getting PSIRT for ' + os_type.upper() +
                 ' device. Version: ' + version)
    ret = s.request('get', url + os_type, params=params)
    if 'errorCode' in ret.json():
        logging.info(os_type.upper() + ' ' + version +
                     ' ' + ret.json()['errorMessage'])
        exit(0)
    else:
        logging.debug('REST API ANSWER:', ret.json())
        advisories = ret.json()['advisories']
        # generate Excel Workbook
        workbook = Workbook()
        worksheet = workbook.active
        headers = ['Advisory ID ', 'Advisory Title', 'CVSS Score', 'URL']
        worksheet.title = 'PSIRT'
        worksheet.append(headers)
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 130
        worksheet.column_dimensions['C'].width = 20
        worksheet.column_dimensions['D'].width = 200
        worksheet['A1'].fill = PatternFill(fgColor=GREEN, fill_type="solid")
        worksheet['B1'].fill = PatternFill(fgColor=GREEN, fill_type="solid")
        worksheet['C1'].fill = PatternFill(fgColor=GREEN, fill_type="solid")
        worksheet['D1'].fill = PatternFill(fgColor=GREEN, fill_type="solid")
        # Populate Workbook and save
        for advisory in advisories:
            worksheet.append([advisory['advisoryId'], advisory['advisoryTitle'],
                              advisory['cvssBaseScore'], advisory['publicationUrl']])
        filename = host.replace('.', '_') + '.xlsx'
        workbook.save(filename)
        logging.info('File: ' + filename + ' was saved.')


def user_password(username=None):
    if username is None:
        username = input('Username: ')
    try:
        password = getpass.getpass(prompt='Password: ')
    except Exception as error:
        logging.error('ERROR', error)
        exit(1)
    return [username, password]


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--host', type=str, default=None,
                        help='hostname/ip address', required=True)
    parser.add_argument('--user', type=str, default=None, help='Username')
    parser.add_argument('--verbose', action='store_true',
                        help='Enable Verbose Output')
    parser.add_argument('--nxapi', action='store_true',
                        help='Connect Using Cisco NX-OS NXAPI')
    args = parser.parse_args()

    format = "%(asctime)s: %(message)s"
    if args.verbose:
        logging.basicConfig(
            format=format, level=logging.DEBUG, datefmt="%H:%M:%S")
    else:
        logging.basicConfig(
            format=format, level=logging.INFO, datefmt="%H:%M:%S")

    host_validator = re.match(
        r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$", args.host)
    if host_validator:
        host = host_validator.group()
    else:
        logging.error('Invalid IP Address. Please validate your input')
        exit(1)

    username, password = user_password(args.user)

    if args.nxapi:
        version = nxos_rest_version(host, username, password)
        device = 'nxos'
    else:
        device, version = ssh_version(host, username, password)

    token = get_auth_token(config.CLIENT_ID, config.CLIENT_SECRET)
    get_psirt(token, device, version)
